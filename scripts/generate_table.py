import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def group_and_replace(array):
    if len(array) == 0:
        return np.array([])

    # array.sort()  # Сортировка массива

    groups = []  # Список для хранения групп чисел
    current_group = [array[0]]  # Текущая группа

    # Итерация по массиву для группировки чисел
    for i in range(1, len(array)):
        if array[i] - current_group[-1] <= 5:  # Если число принадлежит текущей группе
            current_group.append(array[i])
        else:  # Если число не принадлежит текущей группе
            groups.append(current_group)  # Добавляем текущую группу в список групп
            current_group = [array[i]]  # Создаем новую группу с текущим числом

    groups.append(current_group)  # Добавляем последнюю группу

    column_numbers = []
    alphabet = 'abcdefghijklmnopqrstuvwxyz'
    line_number = 0
    additional = ''
    # Замена чисел в каждой группе на их среднее значение
    for group in groups:
        avg_value = np.mean(group)
        for i in range(len(group)):
            group[i] = avg_value
            if i > len(alphabet)-1:
                additional = alphabet[i//len(alphabet)].upper()
            column_numbers.append(additional+alphabet[i % len(alphabet)].upper()+str(line_number+1))
        line_number += 1

    result = np.concatenate(groups)

    return result, column_numbers


def as_text(value):
    if value is None:
        return ""
    return str(value)


def generate_table(cells, number):
    df = pd.DataFrame(data=cells, index=None, columns=["X1", "Y1", "X2", "Y2", "Text"])

    df = df.sort_values(by='Y1')
    Y1, Сell = group_and_replace(df['Y1'].values)
    df['Y1'] = Y1
    df = df.sort_values(by=['Y1','X1'])
    df['Сell'] = Сell
    excel_df = df.drop(['X1', 'Y1', 'X2', 'Y2'], axis=1)

    wb = Workbook()
    ws = wb.active

    for row in excel_df.itertuples():
        cell_name = row[2]
        cell_text = row[1][:-1]
        ws[cell_name] = cell_text

    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)+3
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length

    wb.save(f'excel/table{number}.xlsx')
    wb.close()